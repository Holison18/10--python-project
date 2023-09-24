#This is an attendance tracker application written in python programming language
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# loading the excel book
book = openpyxl.load_workbook(r'C:\Users\Algorithms\Desktop\Holison\Projects\Python\10+ python project\attendanceTracker\data.xlsx')

# choose the sheet
sheet = book['Sheet1']

# counting number of rows/students
row = sheet.max_row

# variable for looping for inputs
response = 1

# counting the number of columns for subjects
columns  = sheet.max_column

# create a list of students to remind
list_of_studentsTR = []

# to concatenate list of roll numbers with lack of attendance
list2 = []

# list of roll numbers with lack of attendance
listofLack = []

# staff mail id
staff_mail = ['kobinaholison2002@gmail.com']

# Warning messages
m1 = "warning!!! you can take only one more day leave for COE 354 class"
m2 = "warning!!! you can take only one more day leave for COE 321 class"
m3 = "warning!!! you can take only one more day leave for COE 356 class"

# define a function to save the excel sheet
def savefile():
    book.save(r'C:\Users\Algorithms\Desktop\Holison\Projects\Python\10+ python project\attendanceTracker\data.xlsx')
    print("Saved!")

# create a function to take user password
def get_email_password():
    password = input("Enter email password: ")
    return password

# create a function to mail student
def mailStudent(listofmails,msg):
    from_id = "kobinaakofiholison@gmail.com"
    password = get_email_password() # get password from user

    # use smtplib to login to the staff email
    s = smtplib.SMTP('smtp.gmail.com',587)
    s.starttls()
    s.login(from_id,password)

    # for each student to warn send an email
    for email in range(0,len(listofmails)):
        to_id = listofmails[email] 
        message = MIMEMultipart()
        message['Subject'] = 'Attendance Report'
        message.attach(MIMEText(msg,'plain'))
        content = message.as_string()
        s.sendmail(from_id,to_id,content)
        s.quit()
    print("Message sent successfully!")

# create a function to send staff a message
def staffMail(mail_id,msg):
    from_id = 'kobinaakofiholison@gmail.com'
    password = get_email_password()

    # use smtplib to login into email
    s = smtplib.SMTP('smtp.gmail.com',587)
    s.starttls() # start ttls
    s.login(from_id,password) # login to the email of the sender

    # send message to staff
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report' # write mail subject
    message.attach(MIMEText(msg,'plain')) # add message attachment
    content = message.as_string()

    # to id
    to_id = mail_id
    s.sendmail(from_id,to_id,content) # send message from sender to the staff
    s.quit() # quit smtp
    print("Staff mail send successfully!")

# create a function to track attendance
def track(no_of_days,row_num,b):

    global staff_mail
    global list2
    global listofLack

    # loop through the list students row_num
    for student in range(0,len(row_num)):
        # if total of number of leaves == warning threshol
        if no_of_days[student] == 2:
            if b == 1:
                list_of_studentsTR.append(sheet.cell(row=row_num[student],column=2).value)
                mailStudent(list_of_studentsTR,m1)
            elif b == 2:
                list_of_studentsTR.append(sheet.cell(row=row_num[student],column=2).value)
                mailStudent(list_of_studentsTR,m2)
            else:
                list_of_studentsTR.append(sheet.cell(row=row_num[student],column=2).value)
                mailStudent(list_of_studentsTR,m3)
        elif no_of_days[student] > 2:
            if b == 1:
                # adding roll no
                list2 = list2+str(sheet.cell(row=row_num[student], column=1).value)
  
                # student mail_id appending
                listofLack.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "COE 354"  # subject based on the code number
  
            elif b == 2:
                list2 = list2+str(sheet.cell(row=row_num[student], column=1).value)
                listofLack.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "COE 321"
  
            else:
                list2 = list2+str(sheet.cell(row=row_num[student], column=1).value)
                listofLack.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "COE 356"
        
        # if treshold is crossed modify the message
        if list2 != "" and len(listofLack) != 0:
            # message for student
            msg1 = "you have lack of attendance in " + subject + " !!!"
  
            # message for staff
            msg2 = "the following students have lack of attendance in your subject : "+list2
  
            mailStudent(listofLack, msg1)  # mail to students
            staff_id = staff_mail  
            staffMail(staff_id, msg2)  # mail to staff

# define a start function which starts the program
def start():
    while True:
        try:
            print("\n\t\t\tAttendance Tracker\n")
            print("Enter the subject number to track attendance")
            print("1. COE 354")
            print("2. COE 321")
            print("3. COE 356")
            print("4. Exit")
            choice = int(input("Enter your choice: "))
            
            # Check if choice is valid (should be between 1 and 4)
            if 1 <= choice <= 4:
                # Input the number of absentees for that subject
                no_of_absentees = int(input('Number of absentees: '))
            
                if no_of_absentees > 1:
                    # If there are multiple absentees, input their roll numbers separated by spaces
                    x = list(map(int, input('Enter roll nos separated by spaces: ').split()))
                else:
                    # If there's only one absentee, input their roll number
                    x = [int(input('Enter roll no: '))]
            
                # Lists to hold row numbers of the students in Excel sheet and the total number of leaves taken
                row_num = []
                no_of_days = []
            
                for student in x:
                    for i in range(2, row+1):
                        if choice == 1:
                            if sheet.cell(row=i, column=1).value == student:
                                m = sheet.cell(row=i, column=3).value
                                m += 1
                                sheet.cell(row=i, column=3).value = m
                                savefile()
                                no_of_days.append(m)
                                row_num.append(i)
                        elif choice == 2:
                            if sheet.cell(row=i, column=1).value == student:
                                m = sheet.cell(row=i, column=4).value
                                m += 1
                                sheet.cell(row=i, column=4).value = m
                                no_of_days.append(m)
                                row_num.append(i)
                        elif choice == 3:
                            if sheet.cell(row=i, column=1).value == student:
                                m = sheet.cell(row=i, column=5).value
                                m += 1
                                sheet.cell(row=i, column=5).value = m
                                row_num.append(i)
                                no_of_days.append(m)
            
                # Call the track function with the collected data
                track(no_of_days, row_num, choice)
            elif choice == 4:
                # Exit the program if choice is 4
                print("Exiting the program.")
                break
            else:
                print("Invalid choice! Please select a valid option (1-4).")
        except ValueError:
            print("Invalid input! Please enter a valid choice.")

if __name__ == '__main__':
    start()